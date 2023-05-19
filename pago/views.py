from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Sum
from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import *
from openpyxl import Workbook
from openpyxl.styles import Font

from .forms import PagoForm
from .models import Pago


class PagoListView(ListView):
    model = Pago
    template_name = 'pago/lista_pago.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Pagos'
        context['create_url'] = reverse_lazy('nuevo_pago')
        context['list_url'] = reverse_lazy('lista_pago')
        context['total_pagadas'] = Pago.objects.filter(estado__contains='PAGADA').count()
        context['total_debe'] = Pago.objects.filter(estado__contains='DEBE').count()
        context['total_anulada'] = Pago.objects.filter(estado__contains='ANULADA').count()
        return context

class PagoCreateView(SuccessMessageMixin, CreateView):
    model = Pago
    template_name = 'pago/nuevo_pago.html'
    form_class = PagoForm
    success_url = reverse_lazy('lista_pago')
    success_message = "%(cliente)s - %(mes)s %(year)s creado exitosamente"

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuevo Pago'
        context['list_url'] = reverse_lazy('lista_pago')
        return context


class PagoUpdateView(SuccessMessageMixin, UpdateView):
    model = Pago
    template_name = 'pago/nuevo_pago.html'
    form_class = PagoForm
    success_url = reverse_lazy('lista_pago')
    success_message = "%(cliente)s - %(mes)s %(year)s editado exitosamente"

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Editar Pago'
        context['list_url'] = reverse_lazy('lista_pago')
        return context


class PagoDetailView(DetailView):
    model = Pago
    template_name = 'pago/detalle_pago.html'

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Detalle Pago'
        context['list_url'] = reverse_lazy('lista_pago')
        return context


# Reporte Excel de Compra
def reporte_excel_pago(request):
    cbo_mes = request.GET['cbo-mes']
    txt_year = request.GET['txt-year']
    cbo_estado = request.GET['cbo-estado']
    txt_contador = request.GET['txt-contador']
    txt_ahorro_vps = request.GET['txt-ahorro-vps']
    txt_sueldo_roraima = request.GET['txt-sueldo-roraima']

    pagos = Pago.objects.all()

    if cbo_mes:
        pagos = pagos.filter(mes__icontains=cbo_mes)
    if txt_year:
        pagos = pagos.filter(year__icontains=txt_year)
    if cbo_estado:
        pagos = pagos.filter(estado__icontains=cbo_estado)

    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    # cabecera
    ws['A1'] = 'MES'
    ws['B1'] = 'AÑO'
    ws['C1'] = 'NOMBRE ÓPTICA'
    ws['D1'] = 'SUBTOTAL'
    ws['E1'] = 'IVA'
    ws['F1'] = 'TOTAL'
    ws['G1'] = 'ESTADO'
    ws['H1'] = 'NRO FACTURA'
    ws['I1'] = 'FECHA PAGO'

    # Iniciar en la 2da fila
    cont = 2
    # Iniciar subtotal general
    subtotal_general = 0

    # Contenido
    for pago in pagos:
        ws.cell(row=cont, column=1).value = pago.mes
        ws.cell(row=cont, column=2).value = pago.year
        ws.cell(row=cont, column=3).value = pago.cliente.nombre_fantasia
        ws.cell(row=cont, column=4).value = pago.subtotal
        ws.cell(row=cont, column=5).value = pago.iva_total
        ws.cell(row=cont, column=6).value = pago.total
        ws.cell(row=cont, column=7).value = pago.estado
        ws.cell(row=cont, column=8).value = pago.n_factura
        ws.cell(row=cont, column=9).value = pago.fecha_factura
        cont += 1

        # calculos
        subtotal_general += pago.subtotal
        total_general = int(subtotal_general) - (int(txt_contador)+int(txt_ahorro_vps)+int(txt_sueldo_roraima))
        total_mitad = int(total_general)/2

    mis_valores = {
        'SUBTOTAL GENERAL': subtotal_general,
        'RESTAR CONTADOR': int(txt_contador),
        'RESTAR AHORRO VPS': int(txt_ahorro_vps),
        'RESTAR SUELDO RORAIMA': int(txt_sueldo_roraima),
        'TOTAL GENERAL': total_general,
        'TOTAL MITAD': total_mitad,
    }

    for clave, valor in mis_valores.items():
        ws.cell(row=cont, column=3).value = clave
        ws.cell(row=cont, column=3).font = Font(bold=True)
        ws.cell(row=cont, column=4).value = valor
        ws.cell(row=cont, column=4).font = Font(bold=True)
        cont += 1

    name_file = 'ReporteExcelPagos.xlsx'
    response = HttpResponse(content_type='application/ms-excel')
    content = 'attachment; filename = {}'.format(name_file)
    response['Content-Disposition'] = content
    wb.save(response)
    return response